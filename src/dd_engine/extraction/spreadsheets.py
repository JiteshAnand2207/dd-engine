"""XLSX extraction that preserves formulas, caches and workbook structure."""

from __future__ import annotations

import io
import warnings
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles.numbers import is_date_format, is_timedelta_format
from openpyxl.utils import get_column_letter, range_boundaries

from dd_engine.extraction.models import JsonObject, SourceExtraction, make_unit


def _json_value(value: object) -> tuple[object, str]:
    if value is None:
        return None, "null"
    if isinstance(value, bool):
        return value, "boolean"
    if isinstance(value, int):
        return value, "integer"
    if isinstance(value, float):
        return value, "number"
    if isinstance(value, Decimal):
        return str(value), "decimal"
    if isinstance(value, datetime):
        return value.isoformat(), "datetime"
    if isinstance(value, date):
        return value.isoformat(), "date"
    if isinstance(value, time):
        return value.isoformat(), "time"
    if isinstance(value, timedelta):
        return value.total_seconds(), "duration_seconds"
    return str(value), "string"


def _format_category(number_format: str) -> str:
    normalized = number_format.casefold()
    if is_date_format(number_format):
        return "date_or_datetime"
    if is_timedelta_format(number_format):
        return "duration"
    if any(token in normalized for token in ("€", "$", "£", "[$", "currency")):
        return "currency"
    if "%" in normalized:
        return "percentage"
    if normalized in {"general", "@"}:
        return "general" if normalized == "general" else "text"
    return "number_or_custom"


def _hidden_columns(worksheet: Any) -> tuple[set[int], list[str]]:
    hidden: set[int] = set()
    labels: list[str] = []
    for label, dimension in worksheet.column_dimensions.items():
        if not dimension.hidden:
            continue
        start = int(dimension.min or 0)
        end = int(dimension.max or start)
        hidden.update(range(start, end + 1))
        labels.append(
            label if start == end else f"{get_column_letter(start)}:{get_column_letter(end)}"
        )
    return hidden, labels


def _defined_names(workbook: Any) -> list[JsonObject]:
    result: list[JsonObject] = []
    collection = workbook.defined_names
    try:
        values = list(collection.values())
    except AttributeError:
        values = list(collection.definedName)
    for item in values:
        destinations: list[JsonObject] = []
        try:
            destinations = [
                {"range": coordinate, "sheet": sheet_name}
                for sheet_name, coordinate in item.destinations
            ]
        except (AttributeError, TypeError, ValueError):
            destinations = []
        result.append(
            {
                "destinations": destinations,
                "hidden": bool(getattr(item, "hidden", False)),
                "name": str(getattr(item, "name", "")),
                "scope_sheet_index": getattr(item, "localSheetId", None),
                "value": str(getattr(item, "attr_text", "") or ""),
            }
        )
    return sorted(result, key=lambda item: (str(item["name"]).casefold(), str(item["value"])))


def _coordinate_in_range(row: int, column: int, coordinate: str) -> bool:
    if "," in coordinate or " " in coordinate:
        return False
    normalized = coordinate.replace("$", "")
    try:
        min_col, min_row, max_col, max_row = range_boundaries(normalized)
    except (TypeError, ValueError):
        return False
    return min_row <= row <= max_row and min_col <= column <= max_col


def _name_memberships(
    names: list[JsonObject], sheet_name: str, row: int, column: int
) -> list[str]:
    memberships: list[str] = []
    for item in names:
        for destination in item["destinations"]:
            if str(destination["sheet"]) == sheet_name and _coordinate_in_range(
                row, column, str(destination["range"])
            ):
                memberships.append(str(item["name"]))
                break
    return memberships


def _structural_roles(value: object, formula: str | None) -> list[str]:
    roles: list[str] = []
    if isinstance(value, str):
        normalized = value.strip().casefold()
        if normalized.startswith("subtotal"):
            roles.append("subtotal_label")
        elif normalized.startswith("total") or normalized.endswith(" total"):
            roles.append("total_label")
    if formula:
        normalized_formula = formula.upper()
        if "SUBTOTAL(" in normalized_formula:
            roles.append("subtotal_formula")
        elif any(token in normalized_formula for token in ("SUM(", "SUMIF(", "SUMIFS(")):
            roles.append("total_formula")
    return roles


def extract_xlsx(*, payload: bytes, source: JsonObject, run_id: str) -> SourceExtraction:
    """Extract all populated workbook cells without recalculating formulas."""

    parser_warnings: list[str] = []
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        formula_book = load_workbook(
            io.BytesIO(payload),
            data_only=False,
            read_only=False,
            keep_links=False,
        )
        cached_book = load_workbook(
            io.BytesIO(payload),
            data_only=True,
            read_only=False,
            keep_links=False,
        )
    parser_warnings.extend(str(item.message) for item in caught)

    names = _defined_names(formula_book)
    units: list[JsonObject] = []
    sheet_coverage: list[JsonObject] = []
    totals: JsonObject = {
        "spreadsheet_cached_formula_values": 0,
        "spreadsheet_cell_units": 0,
        "spreadsheet_currency_formatted_cells": 0,
        "spreadsheet_date_formatted_cells": 0,
        "spreadsheet_formula_errors": 0,
        "spreadsheet_formulas": 0,
        "spreadsheet_hidden_columns": 0,
        "spreadsheet_hidden_rows": 0,
        "spreadsheet_hidden_sheets": 0,
        "spreadsheet_merged_ranges": 0,
        "spreadsheet_named_ranges": len(names),
        "spreadsheet_sheets": len(formula_book.sheetnames),
        "spreadsheet_subtotal_cells": 0,
        "spreadsheet_total_cells": 0,
        "spreadsheet_visible_sheets": 0,
    }

    for sheet_index, sheet_name in enumerate(formula_book.sheetnames, start=1):
        worksheet = formula_book[sheet_name]
        cached_sheet = cached_book[sheet_name]
        sheet_state = str(worksheet.sheet_state)
        if sheet_state == "visible":
            totals["spreadsheet_visible_sheets"] += 1
        else:
            totals["spreadsheet_hidden_sheets"] += 1
        hidden_columns, hidden_column_labels = _hidden_columns(worksheet)
        hidden_rows = sorted(
            int(index)
            for index, dimension in worksheet.row_dimensions.items()
            if bool(dimension.hidden)
        )
        merged_ranges = sorted(str(item) for item in worksheet.merged_cells.ranges)
        totals["spreadsheet_hidden_columns"] += len(hidden_columns)
        totals["spreadsheet_hidden_rows"] += len(hidden_rows)
        totals["spreadsheet_merged_ranges"] += len(merged_ranges)

        raw_cells = list(getattr(worksheet, "_cells", {}).values())
        cells = sorted(raw_cells, key=lambda item: (int(item.row), int(item.column)))
        sheet_unit_start = len(units)
        sheet_formulas = 0
        sheet_cached = 0
        for cell in cells:
            value = cell.value
            if value is None:
                continue
            coordinate = str(cell.coordinate)
            formula = str(value) if cell.data_type == "f" else None
            cached_value_raw = cached_sheet[coordinate].value if formula else None
            cached_value, cached_type = _json_value(cached_value_raw)
            source_value, source_value_type = _json_value(None if formula else value)
            number_format = str(cell.number_format)
            format_category = _format_category(number_format)
            if format_category == "currency":
                totals["spreadsheet_currency_formatted_cells"] += 1
            if format_category == "date_or_datetime":
                totals["spreadsheet_date_formatted_cells"] += 1
            formula_error = None
            if formula and (
                cached_sheet[coordinate].data_type == "e"
                or (isinstance(cached_value_raw, str) and cached_value_raw.startswith("#"))
            ):
                formula_error = str(cached_value_raw)
                totals["spreadsheet_formula_errors"] += 1
            roles = _structural_roles(value, formula)
            totals["spreadsheet_total_cells"] += sum(
                role in {"total_label", "total_formula"} for role in roles
            )
            totals["spreadsheet_subtotal_cells"] += sum(
                role in {"subtotal_label", "subtotal_formula"} for role in roles
            )
            if formula:
                sheet_formulas += 1
                totals["spreadsheet_formulas"] += 1
                if cached_value_raw is not None:
                    sheet_cached += 1
                    totals["spreadsheet_cached_formula_values"] += 1
            merged_memberships = [
                merged for merged in merged_ranges if _coordinate_in_range(cell.row, cell.column, merged)
            ]
            unit_warnings: list[str] = []
            limitation = None
            if formula and cached_value_raw is None:
                limitation = "source formula has no cached value; no recalculation was performed"
                unit_warnings.append("formula_cached_value_missing")
            units.append(
                make_unit(
                    run_id=run_id,
                    source=source,
                    ordinal=len(units) + 1,
                    unit_type="spreadsheet_cell",
                    locator={
                        "cell": coordinate,
                        "column_hidden": int(cell.column) in hidden_columns,
                        "merged_ranges": merged_memberships,
                        "named_ranges": _name_memberships(
                            names, sheet_name, int(cell.row), int(cell.column)
                        ),
                        "range": coordinate,
                        "row_hidden": int(cell.row) in hidden_rows,
                        "sheet": sheet_name,
                        "sheet_index": sheet_index,
                        "sheet_state": sheet_state,
                        "type": "spreadsheet_cell",
                        "workbook": source["relative_path"],
                    },
                    extraction_method="openpyxl_xlsx",
                    confidence=0.99 if formula is None or cached_value_raw is not None else 0.9,
                    content={
                        "cached_value": cached_value if formula else None,
                        "cached_value_type": cached_type if formula else None,
                        "formula": formula,
                        "formula_error": formula_error,
                        "format_category": format_category,
                        "number_format": number_format,
                        "recomputation_status": "not_performed",
                        "recomputed_analytical_value": None,
                        "source_value": source_value,
                        "source_value_type": source_value_type,
                        "structural_roles": roles,
                    },
                    warnings=unit_warnings,
                    limitation=limitation,
                )
            )

        sheet_coverage.append(
            {
                "cached_formula_values": sheet_cached,
                "cell_units": len(units) - sheet_unit_start,
                "formulas": sheet_formulas,
                "hidden_column_ranges": hidden_column_labels,
                "hidden_rows": hidden_rows,
                "max_column": int(worksheet.max_column),
                "max_row": int(worksheet.max_row),
                "merged_ranges": merged_ranges,
                "sheet": sheet_name,
                "sheet_index": sheet_index,
                "state": sheet_state,
            }
        )

    totals["spreadsheet_cell_units"] = len(units)
    totals["spreadsheet_sheet_coverage"] = sheet_coverage
    totals["spreadsheet_defined_names"] = names
    return SourceExtraction(
        status="successfully_extracted",
        primary_method="openpyxl_xlsx",
        units=units,
        warnings=list(dict.fromkeys(parser_warnings)),
        limitation=(
            "workbook formulas were not recalculated; cached values are preserved as source data"
        ),
        metrics=totals,
    )
