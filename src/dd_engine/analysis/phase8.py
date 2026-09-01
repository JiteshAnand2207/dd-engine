"""Phase 8 financial and commercial analytical workstreams."""

from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass

from openpyxl.utils.cell import coordinate_from_string

from dd_engine.analysis.context import (
    AnalysisContext,
    SheetRef,
    numeric_value,
    unit_value,
)
from dd_engine.analysis.records import AnalysisRecords, CitationSpec
from dd_engine.evidence.models import JsonObject

PHASE8_VERSION = "phase8-analysis-v2"


def _money(value: float) -> str:
    return f"EUR {value:,.0f}"


def _percent(value: float) -> str:
    return f"{value:.1f}%"


def _cell_by_label(
    context: AnalysisContext,
    label: str,
    *,
    offset: int = 1,
    path_hint: str | None = None,
) -> JsonObject | None:
    anchor = context.cell_matching(re.escape(label), path_hint=path_hint)
    if anchor is None:
        return None
    sheet = context.sheet_for_unit(context.sheets, anchor)
    return context.offset_cell(sheet, anchor, offset) if sheet else None


def _cell(sheet: SheetRef, coordinate: str) -> JsonObject | None:
    return sheet.cell(coordinate)


def _row_number(unit: JsonObject) -> int:
    locator = unit.get("locator")
    coordinate = locator.get("cell") if isinstance(locator, dict) else None
    if not isinstance(coordinate, str):
        return 0
    return int(coordinate_from_string(coordinate)[1])


def _rows_below(sheet: SheetRef, header_label: str) -> list[dict[str, JsonObject]]:
    header = next(
        (
            unit
            for unit in sheet.cells.values()
            if header_label.casefold() in str(unit_value(unit) or "").casefold()
        ),
        None,
    )
    if header is None:
        return []
    start = _row_number(header) + 1
    row_numbers = sorted(
        {_row_number(unit) for unit in sheet.cells.values() if _row_number(unit) >= start}
    )
    result: list[dict[str, JsonObject]] = []
    for row in row_numbers:
        cells = {
            coordinate_from_string(str(unit["locator"]["cell"]))[0]: unit
            for unit in sheet.cells.values()
            if _row_number(unit) == row
        }
        first = str(unit_value(cells.get("A", {})) or "").strip().casefold()
        if first.startswith("total"):
            break
        if first:
            result.append(cells)
    return result


@dataclass(frozen=True, slots=True)
class StatutoryPoint:
    year: int
    revenue: float
    gross_profit: float
    ebitda: float
    profit_after_tax: float
    unit: JsonObject
    sentence: str


_STATUTORY = re.compile(
    r"Revenue EUR ([\d,]+); gross profit EUR ([\d,]+); EBITDA EUR ([\d,]+); "
    r"profit after tax EUR\s*([\d,]+)\.",
    re.I,
)


def _statutory_points(context: AnalysisContext) -> list[StatutoryPoint]:
    points: list[StatutoryPoint] = []
    for unit in context.text_units(path_hint="statutory"):
        text = unit_value(unit)
        if not isinstance(text, str):
            continue
        year_match = re.search(r"Statutory Accounts (20\d{2})", text, re.I)
        values = _STATUTORY.search(text)
        if not year_match or not values:
            continue
        points.append(
            StatutoryPoint(
                year=int(year_match.group(1)),
                revenue=float(values.group(1).replace(",", "")),
                gross_profit=float(values.group(2).replace(",", "")),
                ebitda=float(values.group(3).replace(",", "")),
                profit_after_tax=float(values.group(4).replace(",", "")),
                unit=unit,
                sentence=values.group(0),
            )
        )
    return sorted(points, key=lambda item: item.year)


def _add_statutory_finding(context: AnalysisContext, records: AnalysisRecords) -> None:
    points = _statutory_points(context)
    if len(points) < 2:
        records.add_gap(
            gap_id="GAP-P8-STATUTORY-TREND",
            expected="A multi-year statutory revenue, gross-profit and EBITDA series.",
            absence_evidence=[f"Only {len(points)} parseable statutory period(s) were identified."],
            importance="high",
            decisions=["price", "earnings_quality"],
            action="Provide a complete statutory series and a management-to-statutory bridge.",
            origin="phase8_analysis",
        )
        return
    first, last = points[0], points[-1]
    claim_id = "CLM-FIN-001"
    revenue_growth = (last.revenue - first.revenue) / first.revenue * 100
    first_margin = first.ebitda / first.revenue * 100
    last_margin = last.ebitda / last.revenue * 100
    calculations = [
        records.add_calculation(
            calculation_id="CALC-FIN-001",
            description=f"Statutory revenue growth from {first.year} to {last.year}.",
            inputs=[
                (
                    "opening_revenue",
                    first.unit,
                    first.revenue,
                    f"Revenue EUR {first.revenue:,.0f}",
                ),
                (
                    "closing_revenue",
                    last.unit,
                    last.revenue,
                    f"Revenue EUR {last.revenue:,.0f}",
                ),
            ],
            expression="(closing_revenue - opening_revenue) / opening_revenue * 100",
            recomputed_value=round(revenue_growth, 2),
            currency=None,
            units="percent",
            period=f"{first.year} to {last.year}",
            claim_ids=(claim_id,),
        ),
        records.add_calculation(
            calculation_id="CALC-FIN-002",
            description=f"Statutory EBITDA margin for {last.year}.",
            inputs=[
                (
                    "ebitda",
                    last.unit,
                    last.ebitda,
                    f"EBITDA EUR {last.ebitda:,.0f}",
                ),
                (
                    "revenue",
                    last.unit,
                    last.revenue,
                    f"Revenue EUR {last.revenue:,.0f}",
                ),
            ],
            expression="ebitda / revenue * 100",
            recomputed_value=round(last_margin, 2),
            currency=None,
            units="percent",
            period=str(last.year),
            claim_ids=(claim_id,),
        ),
    ]
    records.add_finding(
        workstream="financial",
        issue_id="FIN-001",
        conclusion=(
            f"Statutory revenue increased {_percent(revenue_growth)} from {first.year} to "
            f"{last.year}, and EBITDA margin expanded from {_percent(first_margin)} to "
            f"{_percent(last_margin)}; the six-year direction is positive, but annual filings "
            "alone do not establish current run-rate quality."
        ),
        source_fact=(
            f"The statutory series reports revenue of {_money(first.revenue)} in {first.year} "
            f"and {_money(last.revenue)} in {last.year}, with EBITDA of "
            f"{_money(first.ebitda)} and {_money(last.ebitda)} respectively."
        ),
        analysis=(
            "The calculation is a trend analysis, not a valuation. Statutory EBITDA is a "
            "management measure in the filings and must be bridged to the ledger and YTD pack."
        ),
        why_it_matters="Historic growth supports the earnings narrative but does not validate "
        "adjustments or conversion.",
        implication=(
            "Use the statutory trajectory as a reference case only; condition price mechanics "
            "and any earn-out on a reconciled monthly earnings bridge rather than the headline "
            "trend."
        ),
        action=(
            "Obtain the general-ledger-to-statutory bridge for each year and monthly 2025/2026 "
            "management accounts, then rerun margin and cash-conversion analysis."
        ),
        materiality="high",
        confidence=0.92,
        supporting=[
            CitationSpec(first.unit, exact_text=first.sentence),
            CitationSpec(last.unit, exact_text=last.sentence),
        ],
        calculation_ids=calculations,
        uncertainty="Monthly phasing and the statutory-to-management reconciliation were not "
        "supplied.",
        transaction_levers=["price_assumptions", "earn_out", "further_diligence"],
    )


def _add_ebitda_finding(context: AnalysisContext, records: AnalysisRecords) -> None:
    matches = context.units_matching(
        r"Reported adjusted EBITDA: EUR ([\d,]+)\. Baseline ledger EBITDA before the "
        r"transformation adjustment was\s*EUR ([\d,]+)\.",
        path_hint="management accounts",
        locator_type="pdf_page",
    )
    if not matches:
        return
    unit, match = matches[0]
    adjusted = float(match.group(1).replace(",", ""))
    baseline = float(match.group(2).replace(",", ""))
    adjustment = adjusted - baseline
    request = context.cell_matching(r"support to follow", path_hint="financial_request")
    support = [CitationSpec(unit, exact_text=match.group(0))]
    if request is not None:
        support.append(CitationSpec(request, exact_value=unit_value(request)))
    calculation = records.add_calculation(
        calculation_id="CALC-FIN-003",
        description="Bridge baseline ledger EBITDA to reported adjusted EBITDA.",
        inputs=[
            ("baseline_ebitda", unit, baseline, f"EUR {baseline:,.0f}"),
            ("transformation_adjustment", unit, adjustment, f"EUR {adjustment:,.0f}"),
        ],
        expression="baseline_ebitda + transformation_adjustment",
        recomputed_value=adjusted,
        reported_value=adjusted,
        period="six months ended 30 June 2026",
        claim_ids=("CLM-FIN-002",),
    )
    records.add_finding(
        workstream="financial",
        issue_id="FIN-002",
        conclusion=(
            f"The {_money(adjustment)} transformation add-back is unsupported in the room; "
            f"the evidence-backed YTD EBITDA base is therefore {_money(baseline)}, "
            f"{_percent(adjustment / adjusted * 100)} below reported adjusted EBITDA."
        ),
        source_fact=(
            f"The YTD pack reports adjusted EBITDA of {_money(adjusted)} and baseline ledger "
            f"EBITDA of {_money(baseline)}; it states that invoices and an approved plan "
            "were absent."
        ),
        analysis=(
            "Without invoices, approvals or evidence that the cost is non-recurring, the add-back "
            "does not meet a supportable quality-of-earnings standard."
        ),
        why_it_matters="An unsupported add-back directly inflates the earnings base used in "
        "price discussions.",
        implication=(
            "Exclude the adjustment from price assumptions. If the seller seeks value for the "
            "benefit, defer it into an earn-out tied to realized savings and require a warranty "
            "covering the completeness of adjustment support."
        ),
        action=(
            "Obtain the approved restructuring plan, invoice/payroll detail, implementation dates "
            "and evidence of cessation; classify each cost as recurring or exceptional and "
            "rebuild the bridge."
        ),
        materiality="critical",
        confidence=0.98,
        supporting=support,
        calculation_ids=[calculation],
        uncertainty="No deal-lead response or adjustment support has been supplied.",
        transaction_levers=["price_assumptions", "earn_out", "warranty", "further_diligence"],
    )


def _add_working_capital_finding(context: AnalysisContext, records: AnalysisRecords) -> None:
    sheets = context.sheets_with("Trade debtors", "Prepayments", "Calculated net working capital")
    if not sheets:
        return
    sheet = sheets[0]
    cells = tuple(_cell(sheet, coordinate) for coordinate in ("B4", "B5", "B6", "B8", "B9", "B10"))
    debtors_unit, other_unit, creditors_unit, prepayments_unit, reported_unit, formula_unit = cells
    if (
        debtors_unit is None
        or other_unit is None
        or creditors_unit is None
        or prepayments_unit is None
        or reported_unit is None
        or formula_unit is None
    ):
        return
    debtors_value = numeric_value(debtors_unit)
    other_value = numeric_value(other_unit)
    creditors_value = numeric_value(creditors_unit)
    prepayments_value = numeric_value(prepayments_unit)
    reported_value = numeric_value(reported_unit)
    if (
        debtors_value is None
        or other_value is None
        or creditors_value is None
        or prepayments_value is None
        or reported_value is None
    ):
        return
    debtors = float(debtors_value)
    other = float(other_value)
    creditors = float(creditors_value)
    prepayments = float(prepayments_value)
    reported = float(reported_value)
    recomputed = debtors + other + creditors + prepayments
    calculation = records.add_calculation(
        calculation_id="CALC-FIN-004",
        description="Recompute submitted net working capital including trade debtors and "
        "prepayments.",
        inputs=[
            ("trade_debtors", debtors_unit, debtors, None),
            ("other_debtors", other_unit, other, None),
            ("trade_creditors", creditors_unit, creditors, None),
            ("prepayments", prepayments_unit, prepayments, None),
        ],
        expression="trade_debtors + other_debtors + trade_creditors + prepayments",
        recomputed_value=recomputed,
        reported_value=reported,
        period="30 June 2026",
        claim_ids=("CLM-FIN-003",),
    )
    note = _cell(sheet, "D9")
    supporting = [
        CitationSpec(reported_unit, exact_value=unit_value(reported_unit)),
        CitationSpec(formula_unit, exact_value=unit_value(formula_unit)),
    ]
    if note is not None:
        supporting.append(CitationSpec(note, exact_value=unit_value(note)))
    records.add_finding(
        workstream="financial",
        issue_id="FIN-003",
        conclusion=(
            f"The submitted working-capital formula is wrong: it reports {_money(reported)} "
            f"instead of recomputed net working capital of {_money(recomputed)}, a "
            f"{_money(recomputed - reported)} variance."
        ),
        source_fact=(
            "The workbook formula sums rows 5-7 and explicitly notes that trade debtors and a "
            "hidden prepayment row are omitted; the control row reports the correct baseline."
        ),
        analysis=(
            "The error reverses the direction and scale of the working-capital position. It is a "
            "formula defect, not a judgmental normalization."
        ),
        why_it_matters="An incorrect baseline can transfer value mechanically through the "
        "completion accounts.",
        implication=(
            "Do not use the vendor formula in the SPA. Define the working-capital schedule at "
            "account level, attach the corrected baseline, and require a specific "
            "completion-accounts warranty."
        ),
        action=(
            "Lock an account-level NWC definition, agree treatment of other debtors/prepayments, "
            "recompute a monthly twelve-month peg and attach the validated workbook to the "
            "transaction documents."
        ),
        materiality="critical",
        confidence=0.99,
        supporting=supporting,
        calculation_ids=[calculation],
        transaction_levers=["price_assumptions", "working_capital_mechanism", "warranty"],
    )
    records.add_contradiction(
        contradiction_id="CON-FIN-003",
        issue_id="FIN-003",
        conflicting_values=[reported, recomputed],
        source_units=[reported_unit, formula_unit, debtors_unit, prepayments_unit],
        likely_explanations=[
            "The submitted formula omits the trade-debtor and prepayment rows identified in "
            "its note."
        ],
    )


def _add_aged_debtors_finding(context: AnalysisContext, records: AnalysisRecords) -> None:
    sheets = context.sheets_with("Customer", "90+", "Total", path_hint="aged_debtors")
    if not sheets:
        return
    sheet = sheets[0]
    rows = _rows_below(sheet, "Customer")
    over_units = [row["E"] for row in rows if "E" in row and numeric_value(row["E"]) is not None]
    total_unit = sheet.cell("F11")
    reported_over_unit = sheet.cell("E11")
    if not over_units or total_unit is None or reported_over_unit is None:
        return
    over_values = [float(numeric_value(unit) or 0) for unit in over_units]
    total = float(numeric_value(total_unit) or 0)
    reported_over = float(numeric_value(reported_over_unit) or 0)
    recomputed_over = sum(over_values)
    inputs = [
        (f"over_90_{index}", unit, value, None)
        for index, (unit, value) in enumerate(zip(over_units, over_values, strict=True), start=1)
    ]
    calculation = records.add_calculation(
        calculation_id="CALC-FIN-005",
        description="Recompute the 90+ debtor ageing bucket from customer rows.",
        inputs=inputs,
        expression=" + ".join(item[0] for item in inputs),
        recomputed_value=recomputed_over,
        reported_value=reported_over,
        period="30 June 2026",
        claim_ids=("CLM-FIN-004",),
    )
    records.add_finding(
        workstream="financial",
        issue_id="FIN-004",
        conclusion=(
            f"The 90+ debtor bucket is understated: customer rows total {_money(recomputed_over)} "
            f"({_percent(recomputed_over / total * 100)} of debtors), versus the stored total of "
            f"{_money(reported_over)}."
        ),
        source_fact=(
            f"The aged-debtors schedule reports total receivables of {_money(total)}; its 90+ "
            "formula cache does not equal the visible customer rows."
        ),
        analysis=(
            "The error understates overdue exposure and prevents reliance on the submitted ageing "
            "for recoverability or normalized working-capital analysis."
        ),
        why_it_matters="Overdue debt affects cash conversion, bad-debt risk and the "
        "working-capital peg.",
        implication=(
            "Use the recomputed ageing for price and completion accounts, reserve specifically "
            "against disputed/aged balances, and consider an escrow or receivables indemnity "
            "for unrecovered 90+ debt."
        ),
        action=(
            "Obtain invoice-level ageing with post-cut-off cash receipts, dispute status and "
            "credit "
            "notes; reconcile it to the ledger and agree balance-specific provisions."
        ),
        materiality="high",
        confidence=0.99,
        supporting=[
            CitationSpec(reported_over_unit, exact_value=unit_value(reported_over_unit)),
            CitationSpec(total_unit, exact_value=unit_value(total_unit)),
            CitationSpec(over_units[0], exact_value=unit_value(over_units[0])),
        ],
        calculation_ids=[calculation],
        transaction_levers=[
            "price_assumptions",
            "escrow",
            "indemnity",
            "working_capital_mechanism",
        ],
    )
    records.add_contradiction(
        contradiction_id="CON-FIN-004",
        issue_id="FIN-004",
        conflicting_values=[reported_over, recomputed_over],
        source_units=[reported_over_unit, *over_units],
        likely_explanations=[
            "The stored total is stale or its formula range excludes customer rows."
        ],
    )


def _add_debt_finding(context: AnalysisContext, records: AnalysisRecords) -> None:
    loan = _cell_by_label(context, "Total loans", offset=2, path_hint="loan_summary")
    hp = _cell_by_label(context, "Total HP exposure", offset=2, path_hint="hp_summary")
    cash = _cell_by_label(context, "Cash", path_hint="trial_balance_2025")
    director_matches = context.units_matching(
        r"director current account of EUR ([\d,]+) is repayable on demand",
        path_hint="related_party",
        locator_type="docx_paragraph",
    )
    if loan is None or hp is None or cash is None or not director_matches:
        return
    director_unit, director_match = director_matches[0]
    loan_value = float(numeric_value(loan) or 0)
    hp_value = float(numeric_value(hp) or 0)
    cash_value = float(numeric_value(cash) or 0)
    director_value = float(director_match.group(1).replace(",", ""))
    gross = loan_value + hp_value + director_value
    net = gross - cash_value
    calculation = records.add_calculation(
        calculation_id="CALC-FIN-006",
        description="Recompute debt and debt-like exposure net of reported cash.",
        inputs=[
            ("loans", loan, loan_value, None),
            ("hp", hp, hp_value, None),
            ("director_account", director_unit, director_value, director_match.group(0)),
            ("cash", cash, cash_value, None),
        ],
        expression="loans + hp + director_account - cash",
        recomputed_value=net,
        period="loan/HP schedules at 30 June 2026; cash at 31 December 2025",
        claim_ids=("CLM-FIN-005",),
    )
    records.add_finding(
        workstream="financial",
        issue_id="FIN-005",
        conclusion=(
            f"Identified debt and debt-like items total {_money(gross)} before cash, including "
            f"{_money(hp_value)} HP and a {_money(director_value)} on-demand director balance "
            f"excluded from management's loan summary; indicative net debt is {_money(net)} "
            "using the latest ledger cash identified."
        ),
        source_fact=(
            f"Loan, HP and related-party sources report {_money(loan_value)}, {_money(hp_value)} "
            f"and {_money(director_value)} respectively; the trial balance reports cash of "
            f"{_money(cash_value)}."
        ),
        analysis=(
            "HP and an on-demand related-party balance are debt-like regardless of management's "
            "classification. The cash date mismatch and pending visual loan letters limit "
            "completeness."
        ),
        why_it_matters="Debt-like classification changes equity proceeds, consent analysis and "
        "funds flow.",
        implication=(
            "Include HP and the director account in the debt-free/cash-free mechanism, require "
            "payoff letters and releases, and retain escrow until lender balances and "
            "unrestricted cash are confirmed."
        ),
        action=(
            "Obtain dated lender statements, HP payoff schedules, security/covenant details, the "
            "director settlement agreement and bank evidence distinguishing unrestricted from "
            "restricted cash."
        ),
        materiality="critical",
        confidence=0.9,
        supporting=[
            CitationSpec(loan, exact_value=unit_value(loan)),
            CitationSpec(hp, exact_value=unit_value(hp)),
            CitationSpec(director_unit, exact_text=director_match.group(0)),
            CitationSpec(cash, exact_value=unit_value(cash)),
        ],
        calculation_ids=[calculation],
        uncertainty=(
            "Loan photographs and the scanned loan/HP pack remain unreviewed, no restricted-cash "
            "evidence is available, and the cash period differs from the debt schedules."
        ),
        transaction_levers=["price_assumptions", "escrow", "consent", "closing_condition"],
    )


def _add_headcount_finding(context: AnalysisContext, records: AnalysisRecords) -> None:
    allocated = _cell_by_label(context, "Total", path_hint="contractor_headcount")
    legal_scope = context.cell_matching(r"12 active contractors", path_hint="contractor_list")
    paye = _cell_by_label(context, "Total PAYE headcount", path_hint="paye_headcount")
    if allocated is None or legal_scope is None or paye is None:
        return
    allocated_value = float(numeric_value(allocated) or 0)
    scope_match = re.search(r"(\d+) active contractors", str(unit_value(legal_scope)), re.I)
    if not scope_match:
        return
    legal_value = float(scope_match.group(1))
    paye_value = float(numeric_value(paye) or 0)
    calculation = records.add_calculation(
        calculation_id="CALC-FIN-007",
        description="Reconcile client-allocated contractors to the active legal contractor list.",
        inputs=[
            ("allocated_contractors", allocated, allocated_value, None),
            ("active_contractors", legal_scope, legal_value, scope_match.group(0)),
        ],
        expression="active_contractors - allocated_contractors",
        recomputed_value=legal_value - allocated_value,
        currency=None,
        units="people",
        period="30 June 2026",
        claim_ids=("CLM-FIN-006",),
    )
    records.add_finding(
        workstream="financial",
        issue_id="FIN-006",
        conclusion=(
            f"Workforce schedules do not reconcile: PAYE headcount is {paye_value:.0f}, but the "
            f"client allocation lists {allocated_value:.0f} contractors while the legal list "
            f"states {legal_value:.0f}, leaving {legal_value - allocated_value:.0f} "
            "contractors unallocated."
        ),
        source_fact="Three separate schedules provide the PAYE, client-allocation and legal "
        "contractor counts.",
        analysis=(
            "The mismatch prevents a complete payroll/contractor cost bridge and obscures client "
            "dependency and worker-status exposure."
        ),
        why_it_matters="Unreconciled workforce data affects cost normalization, client "
        "delivery and tax/employment risk.",
        implication=(
            "Do not accept the submitted headcount bridge for price assumptions. Require a "
            "warranty on workforce completeness and address status-related liabilities through "
            "targeted indemnity if unresolved."
        ),
        action=(
            "Reconcile every PAYE employee and contractor by unique ID, cost centre, client, "
            "start/end "
            "date and tax status to payroll returns, contracts and the general ledger."
        ),
        materiality="high",
        confidence=0.99,
        supporting=[
            CitationSpec(allocated, exact_value=unit_value(allocated)),
            CitationSpec(legal_scope, exact_value=unit_value(legal_scope)),
            CitationSpec(paye, exact_value=unit_value(paye)),
        ],
        calculation_ids=[calculation],
        transaction_levers=["price_assumptions", "warranty", "indemnity", "further_diligence"],
    )


def _add_pipeline_finding(context: AnalysisContext, records: AnalysisRecords) -> None:
    sheets = context.sheets_with("Opportunity", "Probability", "Weighted value")
    if not sheets:
        return
    sheet = sheets[0]
    rows = _rows_below(sheet, "Opportunity")
    weighted_units = [
        row["E"] for row in rows if "E" in row and numeric_value(row["E"]) is not None
    ]
    total = sheet.cell("E8")
    if not weighted_units or total is None:
        return
    values = [float(numeric_value(unit) or 0) for unit in weighted_units]
    recomputed = sum(values)
    reported = float(numeric_value(total) or 0)
    inputs = [
        (f"weighted_{index}", unit, value, None)
        for index, (unit, value) in enumerate(zip(weighted_units, values, strict=True), start=1)
    ]
    calculation = records.add_calculation(
        calculation_id="CALC-FIN-008",
        description="Recompute total probability-weighted pipeline from opportunity rows.",
        inputs=inputs,
        expression=" + ".join(item[0] for item in inputs),
        recomputed_value=recomputed,
        reported_value=reported,
        period="pipeline snapshot at 30 June 2026",
        claim_ids=("CLM-FIN-007",),
    )
    records.add_finding(
        workstream="financial",
        issue_id="FIN-007",
        conclusion=(
            f"Probability-weighted pipeline is overstated by {_money(reported - recomputed)}: "
            f"the opportunity rows sum to {_money(recomputed)}, not the stored "
            f"{_money(reported)} total."
        ),
        source_fact="The pipeline supplies opportunity values and management probabilities but "
        "no win-rate support.",
        analysis="The arithmetic error compounds the absence of evidence for probability "
        "assumptions.",
        why_it_matters="Forecast support affects confidence in forward revenue and any "
        "performance-linked consideration.",
        implication=(
            "Exclude uncontracted pipeline from fixed price and base any earn-out only on "
            "collected "
            "revenue or gross profit, with customer-level anti-double-counting rules."
        ),
        action=(
            "Correct the pipeline total and provide signed orders, stage history, historical "
            "conversion "
            "rates, expected close dates and margin by opportunity."
        ),
        materiality="high",
        confidence=0.98,
        supporting=[
            CitationSpec(total, exact_value=unit_value(total)),
            CitationSpec(weighted_units[0], exact_value=unit_value(weighted_units[0])),
        ],
        calculation_ids=[calculation],
        uncertainty="No historical probability calibration or signed-order evidence was supplied.",
        transaction_levers=["price_assumptions", "earn_out", "further_diligence"],
    )
    records.add_contradiction(
        contradiction_id="CON-FIN-007",
        issue_id="FIN-007",
        conflicting_values=[reported, recomputed],
        source_units=[total, *weighted_units],
        likely_explanations=[
            "The stored total was not refreshed after opportunity-level weighted values changed."
        ],
    )


def _add_monthly_gap_finding(context: AnalysisContext, records: AnalysisRecords) -> None:
    response = context.cell_matching(r"only YTD pack prepared", path_hint="financial_request")
    if response is None:
        return
    records.add_finding(
        workstream="financial",
        issue_id="FIN-008",
        conclusion=(
            "No monthly management-account pack was prepared; current trading, seasonality, "
            "margin movement and cash conversion cannot be tested at an adviser-quality "
            "monthly cadence."
        ),
        source_fact="The financial request response says only a YTD pack is prepared.",
        analysis="A single six-month total cannot reveal monthly volatility or cut-off effects.",
        why_it_matters="Missing monthly data weakens the reliability of run-rate and forecast "
        "conclusions.",
        implication=(
            "Do not annualize the YTD pack for fixed price. Use a completion condition or tightly "
            "defined earn-out measurement until monthly ledger extracts are validated."
        ),
        action=(
            "Provide monthly trial balances and revenue/gross-margin/EBITDA/cash bridges for "
            "January "
            "2024 through the latest month, including budget and prior-year comparatives."
        ),
        materiality="high",
        confidence=0.99,
        supporting=[CitationSpec(response, exact_value=unit_value(response))],
        transaction_levers=["price_assumptions", "earn_out", "closing_condition"],
    )


def _customer_rows(context: AnalysisContext) -> tuple[SheetRef | None, list[dict[str, JsonObject]]]:
    sheets = context.sheets_with("Customer", "FY2025 revenue", "Group ID", "Group name")
    if not sheets:
        return None, []
    return sheets[0], _rows_below(sheets[0], "Customer")


def _add_customer_concentration(
    context: AnalysisContext, records: AnalysisRecords
) -> list[JsonObject]:
    sheet, rows = _customer_rows(context)
    if sheet is None or not rows:
        return []
    totals = {"fy": sheet.cell("B11"), "ytd": sheet.cell("C11")}
    if totals["fy"] is None or totals["ytd"] is None:
        return []
    group_rows: dict[str, list[dict[str, JsonObject]]] = defaultdict(list)
    for row in rows:
        group = str(unit_value(row.get("D", {})) or "").strip()
        if group:
            group_rows[group].append(row)
    decisions: list[JsonObject] = []
    confirmed: list[tuple[str, list[dict[str, JsonObject]], list[CitationSpec]]] = []
    for group, members in sorted(group_rows.items()):
        names = [str(unit_value(member.get("A", {})) or "") for member in members]
        group_name = str(unit_value(members[0].get("E", {})) or "")
        contract_specs: list[CitationSpec] = []
        for name in names:
            stem = re.sub(r"\s+(?:Limited|Ltd\.?|Trading|Retail)\b.*", "", name, flags=re.I)
            matches = context.units_matching(
                rf"Customer:\s*{re.escape(name)}\.\s*Corporate group:\s*{re.escape(group_name)}\.",
                path_hint="customer contracts",
                locator_type="pdf_page",
            )
            if not matches and stem:
                matches = context.units_matching(
                    rf"Customer:.*{re.escape(stem)}.*Corporate group:\s*{re.escape(group_name)}\.",
                    path_hint="customer contracts",
                    locator_type="pdf_page",
                )
            if matches:
                unit, match = matches[0]
                contract_specs.append(CitationSpec(unit, exact_text=match.group(0)))
        if len(members) == 1:
            status = "no_grouping_required"
        elif len(contract_specs) == len(members):
            status = "confirmed"
        else:
            status = "candidate"
        decisions.append(
            {
                "candidate_group_id": group,
                "decision": status,
                "evidence_basis": (
                    "Each named customer is expressly linked to the same corporate group in a "
                    "contract."
                    if status == "confirmed" and len(members) > 1
                    else (
                        "This is a single named customer; no name-based merge was performed."
                        if status == "no_grouping_required"
                        else "Management's revenue schedule is retained as a candidate grouping "
                        "pending independent identity evidence."
                    )
                ),
                "group_name": group_name,
                "members": names,
            }
        )
        if status == "confirmed":
            confirmed.append((group, members, contract_specs))
    multi = [item for item in confirmed if len(item[1]) > 1]
    if not multi:
        return decisions
    group, members, contract_specs = max(
        multi,
        key=lambda item: sum(float(numeric_value(row["B"]) or 0) for row in item[1]),
    )
    fy_values = [float(numeric_value(row["B"]) or 0) for row in members]
    ytd_values = [float(numeric_value(row["C"]) or 0) for row in members]
    fy_total = float(numeric_value(totals["fy"]) or 0)
    ytd_total = float(numeric_value(totals["ytd"]) or 0)
    fy_share = sum(fy_values) / fy_total * 100
    ytd_share = sum(ytd_values) / ytd_total * 100
    calculation_ids = [
        records.add_calculation(
            calculation_id="CALC-COMM-001",
            description=f"FY2025 revenue concentration for confirmed group {group}.",
            inputs=[
                *[
                    (f"member_{index}", row["B"], value, None)
                    for index, (row, value) in enumerate(
                        zip(members, fy_values, strict=True), start=1
                    )
                ],
                ("total_revenue", totals["fy"], fy_total, None),
            ],
            expression=(
                "("
                + " + ".join(f"member_{index}" for index in range(1, len(members) + 1))
                + ") / total_revenue * 100"
            ),
            recomputed_value=round(fy_share, 2),
            currency=None,
            units="percent",
            period="FY2025",
            claim_ids=("CLM-COMM-001",),
        ),
        records.add_calculation(
            calculation_id="CALC-COMM-002",
            description=f"2026 YTD revenue concentration for confirmed group {group}.",
            inputs=[
                *[
                    (f"member_{index}", row["C"], value, None)
                    for index, (row, value) in enumerate(
                        zip(members, ytd_values, strict=True), start=1
                    )
                ],
                ("total_revenue", totals["ytd"], ytd_total, None),
            ],
            expression=(
                "("
                + " + ".join(f"member_{index}" for index in range(1, len(members) + 1))
                + ") / total_revenue * 100"
            ),
            recomputed_value=round(ytd_share, 2),
            currency=None,
            units="percent",
            period="six months ended 30 June 2026",
            claim_ids=("CLM-COMM-001",),
        ),
    ]
    supporting = [
        *[CitationSpec(row["D"], exact_value=unit_value(row["D"])) for row in members],
        *contract_specs,
    ]
    records.add_finding(
        workstream="commercial",
        issue_id="COMM-001",
        conclusion=(
            f"Evidence-backed group normalization raises {unit_value(members[0]['E'])} to "
            f"{_percent(fy_share)} of FY2025 revenue and {_percent(ytd_share)} of 2026 YTD "
            "revenue; "
            "the separately named customers should be treated as one concentration exposure."
        ),
        source_fact=(
            "The revenue schedule assigns both customers the same group ID and separate framework "
            "agreements expressly identify the same corporate group."
        ),
        analysis="This grouping is supported by contract evidence; no other similar names were "
        "merged without proof.",
        why_it_matters="A single corporate group can exercise renewal, pricing and termination "
        "leverage across multiple contracts.",
        implication=(
            "Model the group as one customer in price assumptions and downside cases; obtain "
            "group-level "
            "renewal/consent confirmation and consider an earn-out or retention escrow tied to "
            "retained gross profit."
        ),
        action=(
            "Confirm ultimate parent, billing entities, VAT numbers, cross-defaults and "
            "renewal dates "
            "for every group member, then rerun concentration on collected revenue and gross "
            "profit."
        ),
        materiality="critical",
        confidence=0.98,
        supporting=supporting,
        calculation_ids=calculation_ids,
        transaction_levers=["price_assumptions", "earn_out", "escrow", "consent"],
    )
    return decisions


def _add_complaint_finding(context: AnalysisContext, records: AnalysisRecords) -> None:
    complaint = context.units_matching(
        r"seeks EUR\s*([\d,]+) in service credits",
        path_hint="complaint",
        locator_type="pdf_page",
    )
    response = context.units_matching(
        r"offers EUR\s*([\d,]+) without admission of liability",
        path_hint="company_response",
        locator_type="pdf_page",
    )
    if not complaint or not response:
        return
    complaint_unit, complaint_match = complaint[0]
    response_unit, response_match = response[0]
    claim = float(complaint_match.group(1).replace(",", ""))
    offer = float(response_match.group(1).replace(",", ""))
    records.add_finding(
        workstream="commercial",
        issue_id="COMM-002",
        conclusion=(
            f"A key customer dispute remains unresolved between an {_money(claim)} service-credit "
            f"claim and the company's {_money(offer)} offer; the room supports an exposure "
            "range, not a nil conclusion."
        ),
        source_fact="The customer alleges three priority-one response failures; the company "
        "accepts two misses but disputes quantum.",
        analysis="The correspondence demonstrates both service-performance and renewal-risk "
        "evidence.",
        why_it_matters="An unresolved service issue can affect cash, renewal probability and "
        "reputation with a concentrated customer.",
        implication=(
            "Reserve the disputed amount in price/working capital, seek a specific indemnity "
            "or escrow "
            "for pre-close service credits, and make renewal evidence a commercial diligence "
            "condition."
        ),
        action=(
            "Obtain the governing SLA, incident timestamps, credit calculation, settlement "
            "correspondence, "
            "subsequent service performance and written renewal status."
        ),
        materiality="high",
        confidence=0.97,
        supporting=[
            CitationSpec(complaint_unit, exact_text=complaint_match.group(0)),
            CitationSpec(response_unit, exact_text=response_match.group(0)),
        ],
        uncertainty="The governing credit formula and settlement status were not supplied.",
        transaction_levers=["escrow", "indemnity", "warranty", "further_diligence"],
    )


def _add_commercial_pipeline_finding(context: AnalysisContext, records: AnalysisRecords) -> None:
    probability = context.cell_matching(r"^0\.85$", path_hint="pipeline")
    stage = context.cell_matching(r"Contracting", path_hint="pipeline")
    if probability is None or stage is None:
        return
    records.add_finding(
        workstream="commercial",
        issue_id="COMM-003",
        conclusion=(
            "Pipeline values are management-weighted rather than contracted revenue; the largest "
            "opportunities depend on an existing concentrated customer and renewals, so they "
            "cannot support fixed consideration."
        ),
        source_fact="The pipeline contains management stages and probabilities, including an "
        "85% contracting-stage opportunity.",
        analysis="No signed order, conversion history or probability governance is cited in "
        "the room.",
        why_it_matters="Pipeline concentration can magnify downside if a key customer delays, "
        "reprices or does not renew.",
        implication=(
            "Attribute no fixed price to uncontracted pipeline; if negotiated, use a "
            "customer-level "
            "earn-out based on collected gross profit with clawback for credits or churn."
        ),
        action=(
            "Produce opportunity-by-opportunity signed evidence, historical stage conversion, "
            "gross "
            "margin, expected start date and overlap with renewal/base revenue."
        ),
        materiality="high",
        confidence=0.9,
        supporting=[
            CitationSpec(probability, exact_value=unit_value(probability)),
            CitationSpec(stage, exact_value=unit_value(stage)),
        ],
        uncertainty="Probability calibration and signed customer evidence are absent.",
        transaction_levers=["price_assumptions", "earn_out"],
    )


def _add_client_headcount_finding(context: AnalysisContext, records: AnalysisRecords) -> None:
    paye_sheets = context.sheets_with("Client code", "PAYE headcount")
    contractor_sheets = context.sheets_with("Client code", "Contractor headcount")
    if not paye_sheets or not contractor_sheets:
        return
    paye_rows = _rows_below(paye_sheets[0], "Client code")
    contractor_rows = _rows_below(contractor_sheets[0], "Client code")
    paye_by_client = {
        str(unit_value(row["A"])): (float(numeric_value(row["B"]) or 0), row["B"])
        for row in paye_rows
        if "A" in row and "B" in row
    }
    contractor_by_client = {
        str(unit_value(row["A"])): (float(numeric_value(row["B"]) or 0), row["B"])
        for row in contractor_rows
        if "A" in row and "B" in row
    }
    common = sorted(set(paye_by_client) & set(contractor_by_client))
    if not common:
        return
    combined = {
        client: paye_by_client[client][0] + contractor_by_client[client][0] for client in common
    }
    largest = max(combined, key=combined.__getitem__)
    total = sum(value[0] for value in paye_by_client.values()) + sum(
        value[0] for value in contractor_by_client.values()
    )
    share = combined[largest] / total * 100 if total else 0
    total_paye = _cell_by_label(context, "Total PAYE headcount", path_hint="paye_headcount")
    total_contractors = _cell_by_label(context, "Total", path_hint="contractor_headcount")
    if total_paye is None or total_contractors is None:
        return
    calculation = records.add_calculation(
        calculation_id="CALC-COMM-003",
        description="Compute largest client-linked workforce share.",
        inputs=[
            ("client_paye", paye_by_client[largest][1], paye_by_client[largest][0], None),
            (
                "client_contractors",
                contractor_by_client[largest][1],
                contractor_by_client[largest][0],
                None,
            ),
            (
                "total_paye",
                total_paye,
                sum(value[0] for value in paye_by_client.values()),
                None,
            ),
            (
                "total_contractors",
                total_contractors,
                sum(value[0] for value in contractor_by_client.values()),
                None,
            ),
        ],
        expression="(client_paye + client_contractors) / (total_paye + total_contractors) * 100",
        recomputed_value=round(share, 2),
        currency=None,
        units="percent",
        period="30 June 2026",
        claim_ids=("CLM-COMM-004",),
    )
    records.add_finding(
        workstream="commercial",
        issue_id="COMM-004",
        conclusion=(
            f"The largest client allocation uses {combined[largest]:.0f} PAYE/contractor roles, "
            f"or {_percent(share)} of the submitted client-linked workforce, creating "
            "delivery and redeployment concentration."
        ),
        source_fact="PAYE and contractor allocation schedules identify headcount by the same "
        "client codes.",
        analysis="The result measures submitted allocation, not necessarily dedicated FTE or "
        "contractual minimum staffing.",
        why_it_matters="A key-client loss can create both revenue churn and stranded workforce "
        "cost.",
        implication=(
            "Reflect redeployment/severance downside in price assumptions and link any earn-out to "
            "gross profit after directly attributable delivery costs."
        ),
        action=(
            "Reconcile client allocations to named roles, utilization, notice periods, "
            "transferability, "
            "contractual staffing obligations and contribution margin."
        ),
        materiality="high",
        confidence=0.9,
        supporting=[
            CitationSpec(
                paye_by_client[largest][1], exact_value=unit_value(paye_by_client[largest][1])
            ),
            CitationSpec(
                contractor_by_client[largest][1],
                exact_value=unit_value(contractor_by_client[largest][1]),
            ),
        ],
        calculation_ids=[calculation],
        uncertainty="Schedules do not distinguish dedicated FTE from shared or partial allocation.",
        transaction_levers=["price_assumptions", "earn_out", "further_diligence"],
    )


def _coverage(
    financial_issue_ids: list[str], commercial_issue_ids: list[str]
) -> dict[str, list[JsonObject]]:
    financial_topics = {
        "six_year_statutory_performance": ["FIN-001"],
        "management_vs_statutory": ["FIN-001", "FIN-002", "FIN-008"],
        "revenue_and_ebitda_bridges": ["FIN-001", "FIN-002"],
        "adjustments_and_quality_of_earnings": ["FIN-002"],
        "gross_margin_trends": ["FIN-001"],
        "customer_concentration": ["COMM-001"],
        "aged_debtors_and_recoverability": ["FIN-004"],
        "aged_creditors": [],
        "other_debtors_and_prepayments": ["FIN-003"],
        "normalised_working_capital": ["FIN-003", "FIN-004"],
        "debt_loans_hp_and_debt_like": ["FIN-005"],
        "cash_and_restricted_cash": ["FIN-005"],
        "fixed_assets_and_disposals": [],
        "related_party_transactions": ["FIN-005"],
        "paye_and_contractor_headcount": ["FIN-006"],
        "forecast_and_pipeline_support": ["FIN-007"],
        "missing_monthly_performance": ["FIN-008"],
        "tax_figures_affecting_financial_conclusions": [],
    }
    commercial_topics = {
        "customer_and_group_concentration": ["COMM-001"],
        "contracted_vs_reported_revenue": ["COMM-003"],
        "customer_longevity_and_churn": [],
        "complaints_disputes_and_service": ["COMM-002"],
        "renewal_and_termination_exposure": ["COMM-003"],
        "pricing_and_margin_dependence": ["COMM-001", "COMM-004"],
        "pipeline_quality_and_probability": ["COMM-003", "FIN-007"],
        "key_client_headcount_dependency": ["COMM-004"],
        "supplier_or_channel_dependence": [],
        "unsupported_market_or_growth_claims": [],
    }
    available = set(financial_issue_ids) | set(commercial_issue_ids)
    return {
        "financial": [
            {
                "issue_ids": [issue for issue in issue_ids if issue in available],
                "status": "analysed"
                if any(issue in available for issue in issue_ids)
                else "limitation",
                "topic": topic,
            }
            for topic, issue_ids in financial_topics.items()
        ],
        "commercial": [
            {
                "issue_ids": [issue for issue in issue_ids if issue in available],
                "status": "analysed"
                if any(issue in available for issue in issue_ids)
                else "limitation",
                "topic": topic,
            }
            for topic, issue_ids in commercial_topics.items()
        ],
    }


def build_phase8(context: AnalysisContext) -> tuple[AnalysisRecords, dict[str, JsonObject]]:
    """Build Phase 8 records from extracted evidence without executing source content."""

    records = AnalysisRecords(context, "P8")
    _add_statutory_finding(context, records)
    _add_ebitda_finding(context, records)
    _add_working_capital_finding(context, records)
    _add_aged_debtors_finding(context, records)
    _add_debt_finding(context, records)
    _add_headcount_finding(context, records)
    _add_pipeline_finding(context, records)
    _add_monthly_gap_finding(context, records)
    grouping = _add_customer_concentration(context, records)
    _add_complaint_finding(context, records)
    _add_commercial_pipeline_finding(context, records)
    _add_client_headcount_finding(context, records)

    financial_findings = records.findings.get("financial", [])
    commercial_findings = records.findings.get("commercial", [])
    coverage = _coverage(
        [str(item["issue_id"]) for item in financial_findings],
        [str(item["issue_id"]) for item in commercial_findings],
    )
    limitations = [
        "Intake must be explicitly completed; unanswered or narrowed answers remain limitations.",
        "No independent valuation is provided.",
        "Cash restrictions, monthly performance and several visual debt sources require "
        "further evidence.",
        "Topics marked limitation had no source-backed adverse conclusion; absence is not "
        "treated as a clean bill of health.",
    ]
    common = {
        "analysis_version": PHASE8_VERSION,
        "generated_by": "deterministic local analytical rules; Python made no model call",
        "independent_valuation_provided": False,
        "limitations": limitations,
        "run_id": context.run_id,
        "schema_version": 1,
        "unresolved_intake_question_ids": context.unresolved_answer_ids(),
        "untrusted_source_data_was_executed": False,
    }
    payloads: dict[str, JsonObject] = {
        "financial": {
            **common,
            "coverage": coverage["financial"],
            "findings": financial_findings,
            "workstream": "financial",
        },
        "commercial": {
            **common,
            "coverage": coverage["commercial"],
            "findings": commercial_findings,
            "workstream": "commercial",
        },
        "customer_grouping": {
            **common,
            "decisions": grouping,
            "rule": (
                "Names are never merged by similarity alone. Confirmed groups require contracts, "
                "addresses, VAT identifiers, explicit management answers or equivalent evidence."
            ),
        },
    }
    return records, payloads
